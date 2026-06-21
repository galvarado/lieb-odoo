"""
Importa el catálogo de puros desde un archivo Excel a Odoo vía XML-RPC.
Idempotente: busca por default_code o barcode antes de crear.

Uso:
    python3 import_catalogo.py \
        --url https://<db>.odoo.com \
        --db <db_name> \
        --apikey <api_key> \
        --excel catalogo_puros.xlsx

Columnas esperadas en el Excel:
    marca, linea, formato, precio, codigo_barras, referencia_interna
"""

import argparse
import sys
import xmlrpc.client

try:
    import openpyxl
except ImportError:
    sys.exit("Instala openpyxl: pip install openpyxl")


CONDICION_ATTR_XMLID = "lieb_puros_heridos.product_attribute_condicion"
CONDICION_LINEA_XMLID = "lieb_puros_heridos.product_attribute_value_linea"
CATEGORIA_PUROS_XMLID = "lieb_puros_heridos.product_category_puros"


def connect(url, db, apikey):
    common = xmlrpc.client.ServerProxy(f"{url}/xmlrpc/2/common")
    uid = common.authenticate(db, "__api__", apikey, {})
    if not uid:
        sys.exit("Autenticación fallida. Verifica URL, base de datos y API key.")
    models = xmlrpc.client.ServerProxy(f"{url}/xmlrpc/2/object")
    return uid, models


def resolve_xmlid(models, url, db, uid, apikey, xmlid):
    module, name = xmlid.split(".")
    result = models.execute_kw(
        db, uid, apikey, "ir.model.data", "search_read",
        [[["module", "=", module], ["name", "=", name]]],
        {"fields": ["res_id"], "limit": 1},
    )
    if not result:
        sys.exit(f"XML ID no encontrado: {xmlid}. Asegúrate de que el módulo esté instalado.")
    return result[0]["res_id"]


def import_catalogo(url, db, uid, apikey, models, excel_path):
    attr_id = resolve_xmlid(models, url, db, uid, apikey, CONDICION_ATTR_XMLID)
    valor_linea_id = resolve_xmlid(models, url, db, uid, apikey, CONDICION_LINEA_XMLID)
    categoria_id = resolve_xmlid(models, url, db, uid, apikey, CATEGORIA_PUROS_XMLID)

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active

    headers = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    required = {"marca", "linea", "formato", "precio"}
    missing = required - set(headers)
    if missing:
        sys.exit(f"Columnas faltantes en el Excel: {missing}")

    creados = omitidos = errores = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))

        if not data.get("marca") or not data.get("linea") or not data.get("formato"):
            continue

        nombre = f"{data['marca']} {data['linea']} {data['formato']}".strip()
        ref = str(data.get("referencia_interna", "")).strip() or None
        barcode = str(data.get("codigo_barras", "")).strip() or None
        precio = float(data.get("precio") or 0)

        domain = []
        if ref:
            domain = [["default_code", "=", ref]]
        elif barcode:
            domain = [["barcode", "=", barcode]]
        else:
            domain = [["name", "=", nombre]]

        existente = models.execute_kw(db, uid, apikey, "product.template", "search", [domain], {"limit": 1})
        if existente:
            omitidos += 1
            continue

        vals = {
            "name": nombre,
            "type": "product",
            "categ_id": categoria_id,
            "list_price": precio,
            "attribute_line_ids": [(0, 0, {
                "attribute_id": attr_id,
                "value_ids": [(4, valor_linea_id)],
            })],
        }
        if ref:
            vals["default_code"] = ref
        if barcode:
            vals["barcode"] = barcode

        try:
            models.execute_kw(db, uid, apikey, "product.template", "create", [vals])
            creados += 1
            print(f"  CREADO: {nombre}")
        except Exception as e:
            errores += 1
            print(f"  ERROR: {nombre} → {e}")

    wb.close()
    print(f"\nResumen: {creados} creados · {omitidos} omitidos · {errores} errores")


def main():
    parser = argparse.ArgumentParser(description="Importa catálogo de puros a Odoo")
    parser.add_argument("--url", required=True, help="URL de la instancia Odoo (ej: https://lieb.odoo.com)")
    parser.add_argument("--db", required=True, help="Nombre de la base de datos")
    parser.add_argument("--apikey", required=True, help="API key del usuario")
    parser.add_argument("--excel", required=True, help="Ruta al archivo Excel del catálogo")
    args = parser.parse_args()

    uid, models = connect(args.url, args.db, args.apikey)
    import_catalogo(args.url, args.db, uid, args.apikey, models, args.excel)


if __name__ == "__main__":
    main()
